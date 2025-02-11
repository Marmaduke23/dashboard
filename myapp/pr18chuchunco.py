import pandas as pd
import requests
from io import BytesIO
from datetime import datetime, time
import os


lista_pozos=["Pozo 1A","Pozo 2A","Pozo 3A","Pozo 4A","Pozo 5"]
planta="San Jose de Chuchunco"
mes=12
año=2024
             
def llenar_excel(df,mes,año,lista_pozos,planta):
    df.columns = df.columns.str.replace('\xa0', ' ', regex=False)

    #Relleno de datos en la planilla

    from openpyxl import load_workbook

    # Obtener la ruta absoluta de la carpeta 'myapp'
    app_path = os.path.dirname(os.path.abspath(__file__))

    # Construir la ruta completa al archivo Excel en la misma carpeta
    file_path = os.path.join(app_path, 'TablaTipo.xlsx')

    #file_path = "TablaTipo.xlsx"
    wb = load_workbook(file_path)
    # Creamos un archivo en memoria (no en disco)
    output = BytesIO()
    for i in lista_pozos:
        print(i.split(" "))
        ws = wb[f"{planta} P{i.split(' ')[1]}"]
        columna_fecha = "A"
        columna_horometro = "B"
        columna_caudal = "D"
        columna_caudal_condicional = "I"
        fila_inicio = 17
        celdames="C14"
        meses_es = [
            "Enero", "Febrero","Marzo","Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
        ]
        mesfin=meses_es[mes-1]
        anofin=año
        Celda= f": {mesfin} {anofin}"
        ws[celdames]=Celda

        #Filtrar datos nan en la columna
        df = df.dropna(subset=[f'{i} - Horómetro'])

        # Extraer solo la fecha (sin la hora)
        df['fecha'] = df['Dif registro'].dt.date
        df['hora'] = df['Dif registro'].dt.time
        df['diff_13h'] = df['hora'].apply(lambda x: abs((datetime.combine(datetime.min, x) - datetime.combine(datetime.min, time(13, 0))).total_seconds()))

        # Definir las fechas de inicio y fin para el filtro
        #fecha de inicio, ultimo dia del mes anterior
        fecha_inicio = datetime(año, mes, 1).date()- pd.DateOffset(days=1)
        #fecha de fin, ultimo dia del mes
        from calendar import monthrange
        ultimo_dia = monthrange(año, mes)[1]  # Devuelve el último día del mes
        fecha_fin = datetime(año, mes, ultimo_dia).date()
        #imprimmimos el nombre del mes de la fecha fin
        mesfin=meses_es[fecha_fin.month-1]
        anofin=fecha_fin.year
        Celda= f": {mesfin} {anofin}"

        # Generar un rango de fechas completas
        rango_fechas = pd.date_range(start=fecha_inicio, end=fecha_fin)

        # Crear un DataFrame con todas las fechas
        df_completo = pd.DataFrame({'fecha': rango_fechas.date})

        # Unir con los datos originales, rellenando los valores NaN con 0
        df_filtrado = df_completo.merge(df, on='fecha', how='left').fillna({f'{i} - Horómetro': 0, 'diff_13h': 99999999999})

        # Aplicar el filtro de fechas
        #df_filtrado = df_final[(df['fecha'] >= fecha_inicio) & (df['fecha'] <= fecha_fin)]

        # Mostrar el resultado filtrado
        df_filtrado = df_filtrado[[f'{i} - Horómetro',f'{i} - Caudal l/s','fecha','diff_13h']]
        # Seleccionar el registro con la menor diferencia por cada día
        df_filtrado = df_filtrado.loc[df_filtrado.groupby('fecha')['diff_13h'].idxmin()]

        # Ordenar por fecha
        df_filtrado = df_filtrado.sort_values(by='fecha')

        #Comprobar si se llena caudal
        valor_caudal = ws["H14"].value
        relleno=False
        if valor_caudal !=0:
            relleno=True


        # Llenamos la tabla en Excel
        for i, row in enumerate(df_filtrado.itertuples(), start=fila_inicio):
            ws[f"{columna_fecha}{i}"] = row.fecha.day  # Inserta la fecha (día)
            ws[f"{columna_horometro}{i}"] = row._1 # Inserta las horas de operación
            if relleno==False:
                ws[f"{columna_caudal_condicional}{i}"] = row._2 
            else: 
                ws[f"{columna_caudal_condicional}{i}"] = valor_caudal               # Inserta las horas de operación
            ws[f"{columna_caudal}{i}"] = row._2          # Inserta las horas de operación

    temp_path = os.path.join(app_path, 'temp.xlsx')        

    # Guardar cambios en un archivo temporal
    wb.save(temp_path)

    # Convertir a BytesIO
    output = BytesIO()
    with open(temp_path, "rb") as f:
        output.write(f.read())
    output.seek(0)

    # Eliminar archivo temporal
    os.remove(temp_path)

    with open("test.xlsx", "wb") as f:
        f.write(output.getvalue())

    return output

#llenar_excel(df,mes,año,lista_pozos,planta)


